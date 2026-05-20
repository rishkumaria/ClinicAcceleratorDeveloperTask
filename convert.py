"""
Excel → JSON Clinic Scoreboard Converter
Author: Rish Kumaria (Rish)
Final Version: 6.0 (The 'Definitive' Edition)

A robust, enterprise-grade extraction tool designed to audit clinical 
performance metrics with zero data loss. This script captures values, 
formulas, hidden states, comments, and semantic units.
"""

from __future__ import annotations

import re
import json
import argparse
from datetime import datetime
from typing import Any, Optional, Dict, List
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

try:
    from pydantic import BaseModel, Field, validator
    PYDANTIC_AVAILABLE = True
except ImportError:
    PYDANTIC_AVAILABLE = False

# ---------------------------------------------------------------------------
# Constants & Header Map
# ---------------------------------------------------------------------------

VERSION = "6.1" # Upgraded with Pydantic and Rule Engine
ROW_SECTION = 1
ROW_METRIC = 2
ROW_FOCUS = 3
ROW_SOURCE = 4
ROW_ROLE = 5
ROW_TARGET_L = 6
ROW_TARGET_V = 7
DATA_START = 8

# --- Pydantic Schemas for Validation ---------------------------------------
if PYDANTIC_AVAILABLE:
    class MetricMetadata(BaseModel):
        name: str
        section: Optional[str]
        focus: Optional[str]
        source: Optional[str]
        role: Optional[str]
        target_type: Optional[str]
        target_value: Optional[Any]
        unit: Optional[str]
        definitions: List[str]
        source_links: List[str]
        excel_info: Dict[str, Any]
        is_calculated: bool

    class Observation(BaseModel):
        date: str
        metric_id: str
        value: Optional[float]
        formula: Optional[str]

    class ScoreboardExport(BaseModel):
        header: Dict[str, Any] = Field(..., alias="_header")
        metrics: Dict[str, MetricMetadata]
        observations: List[Observation]
        rules: List[Dict[str, Any]] # Rule Engine Ingestion

# ---------------------------------------------------------------------------
# Core Converter Logic
# ---------------------------------------------------------------------------

class ScoreboardConverter:
    def __init__(self, input_path: str):
        self.input_path = input_path
        # Pass 1: Results only
        self.ws_data = load_workbook(input_path, data_only=True).active
        # Pass 2: Metadata only (formulas, comments, links, styles)
        self.ws_meta = load_workbook(input_path, data_only=False).active
        
        self.merged_ranges = self.ws_data.merged_cells.ranges
        self.hidden_cols = {
            get_column_letter(i): col.hidden 
            for i, col in enumerate(self.ws_meta.column_dimensions.values(), 1)
        }

    def _get_merged_val(self, row: int, col: int) -> Any:
        """Propagate shared values from merged headers."""
        for mr in self.merged_ranges:
            if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                return self.ws_data.cell(row=mr.min_row, column=mr.min_col).value
        return self.ws_data.cell(row=row, column=col).value

    def _detect_unit(self, fmt: str | None) -> str | None:
        """Heuristically identify the unit/format of a column."""
        if not fmt: return None
        fmt = fmt.lower()
        if "$" in fmt: return "currency"
        if "%" in fmt: return "percent"
        if "h:mm" in fmt: return "time"
        return None

    def _clean(self, val: Any) -> Any:
        """Standardize cell data types for JSON."""
        if val is None: return None
        if isinstance(val, datetime): return val.strftime("%Y-%m-%d")
        if isinstance(val, str):
            if val.startswith("#"): return val  # Error codes
            v = val.strip().replace("\n", " ").replace("\\n", " ")
            v = re.sub(r"\s+", " ", v)
            if v == "": return None
            # Attempt to recover numbers
            try:
                # Handle percentages stored as strings
                if v.endswith("%"): return float(v.strip("%")) / 100.0
                if "." in v: return float(v)
                return int(v)
            except ValueError:
                pass
        return val

    def _slug(self, *parts: str) -> str:
        """Create a clean, unique snake_case identifier."""
        combined = "_".join(str(p) for p in parts if p)
        s = combined.lower().replace("\n", " ").replace("\\n", " ")
        s = re.sub(r"[^a-z0-9]+", "_", s)
        return s.strip("_")

    def parse(self) -> dict[str, Any]:
        metrics_index = {}
        col_to_id = {}
        
        # 1. Schema Analysis
        seen_ids = {}
        for col in range(2, self.ws_data.max_column + 1):
            raw_metric = self.ws_data.cell(row=ROW_METRIC, column=col).value
            
            # Check if this column has ANY data even if Row 2 is empty
            has_observations = any(self.ws_data.cell(row=r, column=col).value is not None 
                                   for r in range(DATA_START, self.ws_data.max_row + 1))
            
            if not raw_metric and not has_observations: 
                continue # Truly a spacer or empty

            col_letter = get_column_letter(col)
            display_name = str(raw_metric).strip() if raw_metric else f"Unlabeled Metric ({col_letter})"
            
            # Contextual metadata
            banner = self._clean(self._get_merged_val(ROW_SECTION, col))
            focus  = self._clean(self.ws_data.cell(row=ROW_FOCUS, column=col).value)
            
            # Construct a unique, namespace-aware ID
            cat_alias = self._slug(banner or focus or "global")
            base_slug = self._slug(display_name)
            unique_id = base_slug if cat_alias in base_slug else f"{cat_alias}_{base_slug}"

            # Deduplication: If ID exists, append column letter to guarantee uniqueness
            if unique_id in seen_ids:
                unique_id = f"{unique_id}_{col_letter.lower()}"
            seen_ids[unique_id] = True
            
            # Lossless Enrichment (Audit all 7 header rows for comments and links)
            all_comments = []
            all_links = []
            data_sample = self.ws_meta.cell(row=DATA_START, column=col)
            
            for h_row in range(1, DATA_START):
                meta_cell = self.ws_meta.cell(row=h_row, column=col)
                
                # 1. Capture Comments
                if meta_cell.comment:
                    all_comments.append(f"Row {h_row}: {meta_cell.comment.text}")
                
                # 2. Capture Hyperlinks (Objects)
                if meta_cell.hyperlink:
                    all_links.append(meta_cell.hyperlink.target)
                
                # 3. Capture Hyperlinks (Formulas)
                elif isinstance(meta_cell.value, str) and "HYPERLINK" in meta_cell.value:
                    match = re.search(r'HYPERLINK\s*\(\s*["\']([^"\']+)["\']', meta_cell.value)
                    if match:
                        all_links.append(match.group(1))

            metrics_index[unique_id] = {
                "name": display_name,
                "section": banner,
                "focus": focus,
                "source": self._clean(self.ws_data.cell(row=ROW_SOURCE, column=col).value),
                "role": self._clean(self.ws_data.cell(row=ROW_ROLE, column=col).value),
                "target_type": self._clean(self.ws_data.cell(row=ROW_TARGET_L, column=col).value),
                "target_value": self._clean(self.ws_data.cell(row=ROW_TARGET_V, column=col).value),
                "unit": self._detect_unit(data_sample.number_format),
                "definitions": all_comments,
                "source_links": list(set(all_links)), # Deduplicated links
                "excel_info": {
                    "column": col_letter,
                    "is_hidden": self.hidden_cols.get(col_letter, False)
                },
                "is_calculated": False # Updated during data pass
            }
            col_to_id[col] = unique_id

        # 2. Timeline Extraction
        timeline = []
        for row in range(DATA_START, self.ws_data.max_row + 1):
            date_cell = self.ws_data.cell(row=row, column=1)
            if not date_cell.value: continue
            
            week_record = {
                "date": self._clean(date_cell.value),
                "observations": {}
            }
            
            for col, mid in col_to_id.items():
                d_cell = self.ws_data.cell(row=row, column=col)
                m_cell = self.ws_meta.cell(row=row, column=col)
                
                val = self._clean(d_cell.value)
                formula = str(m_cell.value) if isinstance(m_cell.value, str) and m_cell.value.startswith("=") else None
                
                if formula:
                    metrics_index[mid]["is_calculated"] = True
                
                if val is not None:
                    week_record["observations"][mid] = {
                        "v": val,
                        "f": formula
                    }
            timeline.append(week_record)

        return {
            "_header": {
                "candidate": "Rish Kumaria",
                "extracted_at": datetime.now().isoformat(),
                "metrics_found": len(metrics_index),
                "weeks_total": len(timeline)
            },
            "metrics": metrics_index,
            "weeks": timeline
        }

    def parse_tidy(self) -> dict[str, Any]:
        """
        Outputs a unified JSON containing:
        1. 'metrics': Full index of all 124+ metrics.
        2. 'observations': Unpivoted Tidy Data records.
        3. 'rules': Rule Engine ingestion (CF Rules).
        """
        full_data = self.parse()
        metrics = full_data["metrics"]
        weeks = full_data["weeks"]
        
        # 1. Unpivot Observations
        tidy_records = []
        for week in weeks:
            date = week["date"]
            for mid, obs in week["observations"].items():
                record = {
                    "date": date,
                    "metric_id": mid,
                    "value": obs["v"],
                    "formula": obs["f"]
                }
                tidy_records.append(record)
        
        # 2. Rule Engine Ingestion (Extract CF Rules)
        rules_export = []
        for cf in self.ws_data.conditional_formatting:
            for rule in cf.rules:
                rules_export.append({
                    "range": str(cf.sqref),
                    "type": rule.type,
                    "operator": getattr(rule, "operator", None),
                    "formula": rule.formula
                })

        output_dict = {
            "_header": full_data["_header"],
            "metrics": metrics,
            "observations": tidy_records,
            "rules": rules_export
        }

        # 3. Schema Validation
        if PYDANTIC_AVAILABLE:
            try:
                # This ensures the dictionary matches our strict professional standards
                validated = ScoreboardExport(**output_dict)
                return validated.dict(by_alias=True)
            except Exception as e:
                print(f"[WARNING] Schema validation failed: {e}")
        
        return output_dict

# ---------------------------------------------------------------------------
# Application Entry Point
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Professional Clinic Scoreboard Converter")
    parser.add_argument("--input", default="Scoreboard Test.xlsx", help="Input Excel path")
    parser.add_argument("--output", default="output.json", help="Output JSON path")
    parser.add_argument("--tidy", action="store_true", help="Output in unpivoted Tidy Data format")
    args = parser.parse_args()

    try:
        converter = ScoreboardConverter(args.input)
        
        if args.tidy:
            result = converter.parse_tidy()
            print(f"[INFO] Using unpivoted Tidy Data format.")
        else:
            result = converter.parse()
            print(f"[INFO] Using Relational Timeline format.")
        
        with open(args.output, "w", encoding="utf-8") as f:
            json.dump(result, f, indent=2, ensure_ascii=False)
            
        print(f"[OK] Conversion complete. {args.output} successfully generated.")
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"[ERR] Fatal error during conversion: {e}")

if __name__ == "__main__":
    main()
